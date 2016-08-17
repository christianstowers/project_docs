#REF: https://automatetheboringstuff.com/chapter12/
#Create a .xlsx file called 'example.xslx' and put it in your working directory

#Import the OpenPyXL module
import openpyxl


#Use the openpyxl.load_workbook() function to load a workbook
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
# <class 'openpyxl.workbook.workbook.Workbook'>


#Use get_sheet_names() to get a list of all the sheet names in a workbook
wb.get_sheet_names()
# ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb.get_sheet_by_name('Sheet3')
sheet
# <Worksheet "Sheet3">
type(sheet) 
# <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title
# 'Sheet3'
anotherSheet = wb.active
anotherSheet
# <Worksheet "Sheet1">


#With access to a Worksheet object, you can access a Cell object by its name.
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1']
# <Cell Sheet1.A1>
sheet['A1'].value
#datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']
c.value
# 'Apples'
'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
# 'Row 1, Column B is Apples'
'Cell ' + c.coordinate + ' is ' + c.value
# 'Cell B1 is Apples'
sheet['C1'].value
# 73


#Accessing the cell() method to access cell data by 
#passing integers for row and column keyword arguments
sheet.cell(row=1, column=2)
# <Cell Sheet1.B1>
sheet.cell(row=1, column=2).value
# 'Apples'
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
'''
1 Apples
3 Pears
5 Apples
7 Strawberries
'''